from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from io import BytesIO
from collections import defaultdict
import logging

logger = logging.getLogger("workbook_writer")


class ReconciliationWorkbookWriter:

    def __init__(self, business_date: str):
        """business_date format: YYYY-MM-DD"""
        self.business_date = business_date
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def get_filename(self) -> str:
        date_str = datetime.strptime(self.business_date, "%Y-%m-%d").strftime("%m-%d-%Y")
        return f"Reconciliation_{date_str}.xlsx"

    # ================= SHEET 1: CXP WITH HIGHLIGHTING =================
    def create_cxp_sheet(
        self,
        sales_orders: list,
        order_items: list,
        classification: dict,
        converge_current: dict,
        converge_settled: dict,
        asn_process_numbers: list
    ):
        sheet = self.workbook.create_sheet("CXP")

        headers = [
            "Order number", "Email", "Date", "CXP DB Status",
            "Phone number", "Payment Reference", "CXP Status",
            "Converge Status", "Converge Settled", "ASN"
        ]
        sheet.append(headers)
        self._apply_header_style(sheet, 1, len(headers))

        customer_history = defaultdict(list)
        for order in sales_orders:
            key = order.get("notif_email") or order.get("notify_mobile_no")
            if key:
                customer_history[key].append(order["process_number"])

        multi_retry_orders = {
            oid
            for orders in customer_history.values()
            if len(orders) > 1
            for oid in orders
        }

        converge_invoices = converge_current.get("invoices", {})
        settled_invoices  = converge_settled.get("invoice_level", {})
        asn_set           = set(asn_process_numbers)

        row_num = 2
        for order in sales_orders:
            order_id      = order["process_number"]
            order_class   = classification["orders"].get(order_id, {})
            converge_info = converge_invoices.get(order_id, {})
            settled_info  = settled_invoices.get(order_id, {})

            converge_status = converge_info.get("final_status", "NA")
            is_settled      = "Yes" if settled_info.get("settled", False) else "No"
            asn_status      = "1" if order_id in asn_set else ""

            sheet.append([
                order_id,
                order.get("notif_email"),
                order.get("order_date"),
                order.get("order_state"),
                order.get("notify_mobile_no"),
                order.get("payment_reference_no"),
                order.get("fulfillment_status"),
                converge_status,
                is_settled,
                asn_status
            ])

            if order_id in multi_retry_orders:
                state = order_class.get("state")
                colour = "C6EFCE" if state == "SUCCESS" else ("FFEB9C" if state == "FAILED" else "FFF2CC")
                self._apply_row_fill(sheet, row_num, len(headers), colour)

            row_num += 1

        self._apply_borders(sheet)
        self._auto_fit_columns(sheet)

    # ================= SHEET 2: CONVERGE WITH HIGHLIGHTING =================
    def create_converge_sheet(self, converge_rows: list, converge_current: dict, sales_orders: list):
        sheet = self.workbook.create_sheet("Converge")

        headers = [
            "Invoice Number", "Auth Message", "customer Full Name",
            "Transaction Date", "CXP DB Status", "Email",
            "Card Related Issues", "For Multiple Tries"
        ]
        sheet.append(headers)
        self._apply_header_style(sheet, 1, len(headers))

        cxp_lookup = {
            o["process_number"]: {
                "cxp_db_status": o.get("order_state", ""),
                "email":         o.get("notif_email", "")
            }
            for o in sales_orders
        }

        converge_invoices = converge_current.get("invoices", {})

        row_num = 2
        for row in converge_rows:
            invoice      = (row.get("Invoice Number") or "").strip()
            auth_message = (row.get("Auth Message") or "").strip()

            is_card_issue = any(
                kw in auth_message.upper()
                for kw in ("DECLINED", "SUSPECTED FRAUD", "NSF", "CLOSED", "WITHDRAWAL")
            )

            invoice_data   = converge_invoices.get(invoice, {})
            is_data_issue  = invoice_data.get("is_data_issue", False)
            multiple_tries = "Yes" if is_data_issue else ""

            cxp_data = cxp_lookup.get(invoice, {})

            sheet.append([
                invoice,
                auth_message,
                row.get("Customer Full Name"),
                row.get("Transaction Date"),
                cxp_data.get("cxp_db_status", ""),
                cxp_data.get("email", ""),
                "Yes" if is_card_issue else "",
                multiple_tries
            ])

            if is_card_issue:
                self._apply_row_fill(sheet, row_num, len(headers), "DAEEF3")

            row_num += 1

        self._apply_borders(sheet)
        self._auto_fit_columns(sheet)

    # ================= SHEET 3: CONVERGE SETTLED WITH HIGHLIGHTING =================
    def create_converge_settled_sheet(self, settled_rows: list):
        sheet = self.workbook.create_sheet("Converge Settled")

        headers = [
            "INVOICE NUMBER", "AMOUNT", "Transaction Status", "Original Transaction Type"
        ]
        sheet.append(headers)
        self._apply_header_style(sheet, 1, len(headers))

        row_num = 2
        for row in settled_rows:
            invoice = (row.get("Invoice Number") or "").strip()
            if not invoice:
                continue

            transaction_status = (row.get("Transaction Status") or "").strip()
            is_discrepancy     = transaction_status.upper() != "SETTLED"

            sheet.append([
                invoice,
                row.get("Original Amount"),
                transaction_status,
                row.get("Original Transaction Type")
            ])

            if is_discrepancy:
                self._apply_row_fill(sheet, row_num, len(headers), "FFC7CE")

            row_num += 1

        self._apply_borders(sheet)
        self._auto_fit_columns(sheet)

    # ================= SHEET 4: ORDERS SHIPPED =================
    def create_orders_shipped_sheet(
        self,
        asn_process_numbers: list,
        asn_order_totals: list,
        converge_settled: dict,
        classification: dict
    ):
        """
        Flow:
          Query 3 → fetch_asn_process_numbers  → asn_process_numbers  → Column A / E
          Query 4 → fetch_order_totals(ASN ids) → asn_order_totals     → Column B
          Converge Settled batch                                         → Column F
          Compare B vs F:
            - Match (diff <= 0.01)  → Differences = "NO",  no highlight
            - Mismatch (diff > 0.01)→ Differences = amount, red row → ACTION REQUIRED
            - Not in Converge       → Differences = "NOT IN CONVERGE", red row → ACTION REQUIRED
            - CXP total missing     → Column B blank, yellow row (data gap)
        """
        sheet = self.workbook.create_sheet("Orders Shipped")

        headers = [
            "Order no.", "Sum of Total CXP", "MATCHING WITH CONVERGE",
            "Differences", "Order no", "Amount"
        ]
        sheet.append(headers)
        self._apply_header_style(sheet, 1, len(headers))

        if not asn_process_numbers:
            sheet.append(["No ASN orders found for this period."])
            self._apply_borders(sheet)
            self._auto_fit_columns(sheet)
            return

        # Build CXP order total map from Query 4 results.
        # Key = process_number (normalised), Value = order_total from DB.
        # asn_order_totals is fetch_order_totals(asn_process_numbers) — ASN ids only.
        order_total_map = {}
        for item in (asn_order_totals or []):
            pnum  = item.get("process_number")
            total = item.get("order_total")
            if pnum and total is not None:
                order_total_map[str(pnum).strip()] = total

        settled_invoices = converge_settled.get("invoice_level", {})

        # Settled amount — first SALE row per invoice (user confirmed correct)
        settled_amount_map = {}
        for invoice, data in settled_invoices.items():
            for raw_row in data.get("raw_rows", []):
                if raw_row.get("transaction_type") == "SALE":
                    settled_amount_map[invoice] = raw_row.get("amount")
                    break

        # Build action_required set for quick lookup (already flagged by classifier)
        action_required_set = set(classification.get("action_required_orders", []))

        row_num = 2
        for order_id in asn_process_numbers:
            oid_str         = str(order_id).strip()
            cxp_amount      = order_total_map.get(oid_str)      # Column B — from DB (Query 4)
            converge_amount = settled_amount_map.get(oid_str)   # Column F — from Converge settled
            is_in_converge  = oid_str in settled_invoices
            matching_status = "YES" if is_in_converge else "NO"

            difference = ""
            needs_action = oid_str in action_required_set

            if not is_in_converge:
                # Shipped but not in Converge settled — critical gap
                difference   = "NOT IN CONVERGE"
                needs_action = True
            elif cxp_amount is not None and converge_amount is not None:
                try:
                    diff_val = abs(float(cxp_amount) - float(converge_amount))
                    if diff_val > 0.01:
                        difference   = f"{diff_val:.2f}"
                        needs_action = True
                    else:
                        difference = "NO"   # amounts match — no issue
                except (TypeError, ValueError):
                    difference = "ERROR"
            elif cxp_amount is None and is_in_converge:
                # In Converge but DB total missing — data gap, yellow highlight
                difference = ""

            sheet.append([
                oid_str,
                cxp_amount,          # blank if DB total not found — shows gap clearly
                matching_status,
                difference,
                oid_str,
                converge_amount if converge_amount is not None else "NA"
            ])

            if not is_in_converge or (difference and difference not in ("NO", "")):
                # Red = action required (missing from Converge or amount mismatch)
                self._apply_row_fill(sheet, row_num, len(headers), "FFC7CE")
                if not is_in_converge:
                    try:
                        sheet.cell(row=row_num, column=3).comment = Comment(
                            "Shipped (ASN received) but payment not settled in Converge",
                            "Reconciliation"
                        )
                    except Exception:
                        pass
            elif cxp_amount is None and is_in_converge:
                # Yellow = data gap — in Converge but CXP total missing from DB
                self._apply_row_fill(sheet, row_num, len(headers), "FFEB9C")

            row_num += 1

        self._apply_borders(sheet)
        self._auto_fit_columns(sheet)

    # ================= SHEET 5: RECONCILIATION =================
    def create_reconciliation_sheet(
        self,
        cxp_orders: list,
        classification: dict,
        converge_current: dict,
        converge_settled: dict,
        asn_process_numbers: list,
        order_totals: list,
        asn_order_totals: list = None
    ):
        """
        6-section reconciliation report. Uses explicit sheet.cell(row, col)
        throughout — never sheet.append() — so every header lands on exactly
        the right row with no drift.
        """
        sheet = self.workbook.create_sheet("Reconciliation", 0)
        r = 1   # single source of truth — equals actual sheet row

        converge_invoices = converge_current.get("invoices", {})
        settled_invoices  = converge_settled.get("invoice_level", {})
        asn_set           = set(asn_process_numbers)
        orders_dict       = classification["orders"]

        order_total_map = {}
        for item in (order_totals or []):
            pnum  = item.get("process_number")
            total = item.get("order_total")
            if pnum and total is not None:
                order_total_map[str(pnum).strip()] = total

        # asn_total_map: order totals fetched specifically for ASN orders (Q4).
        # These are fetched by process_number with no date filter on the ORDER,
        # so they include orders placed outside the current date window (e.g.,
        # ordered Monday, ASN logged Friday). This is the correct amount source
        # for Section 2 — order_total_map only has orders in the CXP date window.
        asn_total_map = {}
        for item in (asn_order_totals or []):
            pnum  = item.get("process_number")
            total = item.get("order_total")
            if pnum and total is not None:
                asn_total_map[str(pnum).strip()] = total
        # Fallback: for any ASN order not in asn_total_map, try order_total_map
        for k, v in order_total_map.items():
            if k not in asn_total_map:
                asn_total_map[k] = v

        # ── Helpers ────────────────────────────────────────────────────────────
        def wr(row, values):
            for col, val in enumerate(values, start=1):
                sheet.cell(row=row, column=col).value = val

        def section_hdr(row, title):
            sheet.cell(row=row, column=1).value = title
            self._apply_header_style(sheet, row, 6)
            sheet.merge_cells(
                start_row=row, start_column=1,
                end_row=row,   end_column=6
            )
            return row + 1

        def sub_hdr(row, cols):
            wr(row, cols)
            self._apply_sub_header_style(sheet, row, len(cols))
            return row + 1

        def no_issue_row(row):
            sheet.cell(row=row, column=4).value = "No issue"
            return row + 1

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 1 — Orders present in CXP but NOT in Converge
        # Scope: fulfillment_status ORDERED or CLAIMED — payment expected but
        #        zero presence in either Converge batch.
        # SHIPPED orders not in Converge settled go to Section 4 (ASN_NOT_SETTLED).
        # REJECTED: new scenario — included for visibility; a rejected order may
        #           legitimately have no Converge record (also logged in Logs sheet).
        # Columns: Order # | Converge Status | CXP Status | Remarks
        # ══════════════════════════════════════════════════════════════════════
        r = section_hdr(r, "Orders present in CXP and not present in Converge")
        r = sub_hdr(r, ["Order #", "Converge Status", "CXP Status", "Remarks"])

        _sec1_statuses = {"ORDERED", "CLAIMED", "REJECTED"}
        sec1_orders = [
            o for o in cxp_orders
            if o.get("fulfillment_status") in _sec1_statuses
            and o["process_number"] not in converge_invoices
            and o["process_number"] not in settled_invoices
        ]
        if sec1_orders:
            for order in sec1_orders:
                fs     = order.get("fulfillment_status", "")
                remark = (
                    "REJECTED — may have no payment record; also logged in Logs sheet"
                    if fs == "REJECTED" else ""
                )
                wr(r, [order["process_number"], "NA", fs, remark])
                colour = "FFC7CE" if fs == "REJECTED" else "FFEB9C"
                self._apply_row_fill(sheet, r, 4, colour)
                r += 1
        else:
            r = no_issue_row(r)

        r += 2

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 2 — Orders Shipped in CXP but amount different in Converge
        # Columns: Order # | Converge Amount | CXP Amount | Remarks
        # ══════════════════════════════════════════════════════════════════════
        r = section_hdr(r, "Orders Shipped in CXP but amount different in Converge")
        r = sub_hdr(r, ["Order #", "Converge Amount", "CXP Amount", "Remarks"])

        settled_amount_map = {}
        for invoice, data in settled_invoices.items():
            for raw_row in data.get("raw_rows", []):
                if raw_row.get("transaction_type") == "SALE":
                    settled_amount_map[invoice] = raw_row.get("amount")
                    break

        # Loop over ASN orders — not cxp_orders — because an order can be in
        # ASN from a different day's date window (e.g., ordered Monday, ASN
        # logged Friday). cxp_orders only has Friday's orders; asn_process_numbers
        # has everything the warehouse shipped in this ASN window.
        # Use asn_total_map (Q4, no date filter on order) for the CXP amount.
        sec2_written = False
        for oid in asn_process_numbers:
            oid = str(oid).strip()
            cxp_amount      = asn_total_map.get(oid)
            converge_amount = settled_amount_map.get(oid)
            if cxp_amount is not None and converge_amount is not None:
                try:
                    # difference = Converge - CXP
                    # negative → Converge under-settled (needs to collect more)
                    # positive → Converge over-settled (collected too much)
                    diff_val = float(converge_amount) - float(cxp_amount)
                    if abs(diff_val) > 0.01:
                        sign = "+" if diff_val > 0 else ""
                        remark = (
                            f"Difference: {sign}{diff_val:.2f} — "
                            "Settled amount differs from CXP order total; please verify with finance"
                        )
                        wr(r, [oid, converge_amount, cxp_amount, remark])
                        self._apply_row_fill(sheet, r, 4, "FFC7CE")
                        r += 1
                        sec2_written = True
                except (TypeError, ValueError):
                    pass
        if not sec2_written:
            r = no_issue_row(r)

        r += 2

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 3 — Orders Shipped in CXP but NOT showing in ASN
        # Source of truth for "shipped": fulfillment_status == "SHIPPED"
        #   (from order_items DB query — query 2)
        # Source of truth for "ASN received": asn_process_numbers
        #   (from fetch_asn_process_numbers DB query — query 3)
        # SHIPPED ONLY — CLAIMED is intentionally excluded.
        # Columns: Order # | ASN Status | CXP Status | Remarks
        # ══════════════════════════════════════════════════════════════════════
        r = section_hdr(r, "Orders Claimed / Shipped in CXP but not showing in ASN")
        r = sub_hdr(r, ["Order #", "ASN Status", "CXP Status", "Remarks"])

        shipped_not_in_asn = [
            o for o in cxp_orders
            if o.get("fulfillment_status") == "SHIPPED"
            and o["process_number"] not in asn_set
        ]
        if shipped_not_in_asn:
            for order in shipped_not_in_asn:
                wr(r, [
                    order["process_number"],
                    "NA",
                    order.get("fulfillment_status", ""),
                    ""
                ])
                self._apply_row_fill(sheet, r, 4, "FFDCB2")
                r += 1
        else:
            r = no_issue_row(r)

        r += 2

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 4 — Orders Shipped in CXP but Converge is NOT settled
        # ASN_NOT_SETTLED    — ASN exists (warehouse shipped) but no settlement
        # SHIPPED_NOT_SETTLED— CXP says SHIPPED but not settled (no ASN)
        # Both are the same business problem: money not collected for shipped order.
        # Columns: Order # | CXP status | Converge Status | Remarks
        # ══════════════════════════════════════════════════════════════════════
        r = section_hdr(r, "Orders Shipped in CXP but converge is not settled")
        r = sub_hdr(r, ["Order #", "CXP status", "Converge Status", "Remarks"])

        _sec4_reasons = {"ASN_NOT_SETTLED", "SHIPPED_NOT_SETTLED"}
        sec4_written = False
        for order in cxp_orders:
            oid         = order["process_number"]
            order_class = orders_dict.get(oid, {})
            reason      = order_class.get("action_reason", "")
            if reason in _sec4_reasons:
                remark = (
                    "ASN received — warehouse shipped but payment not settled"
                    if reason == "ASN_NOT_SETTLED"
                    else "CXP shipped — no ASN, not settled"
                )
                wr(r, [
                    oid,
                    order_class.get("cxp_status", "NA"),
                    order_class.get("converge_status", "NA"),
                    remark
                ])
                self._apply_row_fill(sheet, r, 4, "FFC7CE")
                r += 1
                sec4_written = True
        if not sec4_written:
            r = no_issue_row(r)

        r += 3

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 5 — Summary Statistics (5 lines)
        # ══════════════════════════════════════════════════════════════════════
        total_orders     = len(cxp_orders)
        successful_count = len(classification["successful_orders"])
        retry_count      = len(classification["retry_success_orders"])

        declined_count = sum(
            1 for oid, data in orders_dict.items()
            if data.get("state") == "FAILED"
            and "DECLINED" in data.get("converge_status", "")
        )
        cancelled_count = sum(
            1 for o in cxp_orders
            if o.get("order_state") == "PAYMENT_CANCELLED"
        )

        summary_data = [
            ("Total Number of orders submitted",                            total_orders),
            ("Number of orders placed successfully",                        successful_count),
            ("Number of orders declined due to credit card related issues", declined_count),
            ("Number of orders cancelled by users (user-initiated)",        cancelled_count),
            ("Number of orders placed successfully after multiple tries",   retry_count),
        ]

        for label, value in summary_data:
            sheet.cell(row=r, column=1).value = label
            sheet.cell(row=r, column=6).value = value
            sheet.cell(row=r, column=6).font  = Font(bold=True)
            r += 1

        r += 2

        # ══════════════════════════════════════════════════════════════════════
        # SECTION 6 — Failed Orders Detail
        # Columns: OrderNumber | Email | Phone Number | CXP Status |
        #          Converge Status | Reason for order failure
        # ══════════════════════════════════════════════════════════════════════
        r = section_hdr(r, "Order Details")
        r = sub_hdr(r, [
            "Order Number", "Email", "Phone Number",
            "CXP Status", "Converge Status", "Reason for order failure"
        ])

        failed_ids     = set(classification["failed_orders"])
        internal_users = {"amit.kumar@phasezeroventures.com","shankar.puppala@phasezero.ai","krishna.majeti@phasezer.ai"}


        previous_failed_ids = set()
        for oid in classification.get("retry_success_orders", []):
            prev = orders_dict.get(oid, {}).get("previous_failed_attempt")
            if prev:
                previous_failed_ids.add(prev)

        for order in cxp_orders:
            oid   = order["process_number"]
            email = (order.get("notif_email") or "").strip().lower()
            if email in internal_users or oid not in failed_ids:
                continue

            oc        = orders_dict.get(oid, {})
            db_status = oc.get("cxp_db_status", "")
            conv_stat = oc.get("converge_status", "NA")

            if db_status == "PAYMENT_CANCELLED":
                reason = "user-initiated cancellation"
            elif "DECLINED" in conv_stat:
                reason = "card declined"
            elif conv_stat == "SUSPECTED FRAUD":
                reason = "suspected fraud"
            else:
                reason = ""

            wr(r, [
                oid,
                order.get("notif_email", ""),
                order.get("notify_mobile_no", ""),
                db_status,
                conv_stat,
                reason
            ])

            sheet.cell(row=r, column=1).alignment = Alignment(horizontal="center")
            sheet.cell(row=r, column=3).alignment = Alignment(horizontal="center")

            if oid in previous_failed_ids:
                self._apply_row_fill(sheet, r, 1, "DAEEF3")
            r += 1

        self._apply_borders(sheet)
        self._auto_fit_columns(sheet)

    # ================= SHEET 6: LOGS =================
    def create_logs_sheet(
        self,
        start_date: str,
        end_date: str,
        sales_orders: list,
        order_items: list,
        asn_process_numbers: list,
        asn_order_totals: list,
        order_totals: list,
        classification: dict,
        converge_current_result: dict,
        converge_settled_result: dict
    ):
        """
        All information that would go to the log file — run metadata,
        DB query counts, converge batch stats, full classification breakdown,
        action required order details, settlement mismatches, retry successes.
        """
        sheet = self.workbook.create_sheet("Logs")
        r     = 1

        orders_dict = classification.get("orders", {})
        stats       = classification.get("classification_stats", {})

        # ── Pre-compute: full amount mismatch list from ASN orders ────────────
        # This mirrors Section 2 of the Reconciliation sheet exactly so that
        # the Logs count and detail table always match the Reconciliation sheet.
        # The classifier's settlement_amount_mismatches only covers cxp_orders
        # (date-window orders); here we use all ASN orders (wider scope).
        _asn_settled_invoices = converge_settled_result.get("invoice_level", {})
        _asn_settled_amount_map = {}
        for _inv, _data in _asn_settled_invoices.items():
            for _raw in _data.get("raw_rows", []):
                if _raw.get("transaction_type") == "SALE":
                    _asn_settled_amount_map[_inv] = _raw.get("amount")
                    break

        _asn_total_map = {}
        _asn_order_state_map = {}
        for _item in (order_totals or []):
            _pnum = _item.get("process_number")
            _tot = _item.get("order_total")
            if _pnum and _tot is not None:
                _asn_total_map[str(_pnum).strip()] = _tot
        for _item in (asn_order_totals or []):
            _pnum = _item.get("process_number")
            _tot = _item.get("order_total")
            _state = _item.get("order_state")
            if _pnum:
                _pnum = str(_pnum).strip()
                if _tot is not None:
                    _asn_total_map[_pnum] = _tot
                if _state is not None:
                    _asn_order_state_map[_pnum] = _state

        full_amount_mismatches = []
        for _oid in (asn_process_numbers or []):
            _oid = str(_oid).strip()
            _cxp_amt = _asn_total_map.get(_oid)
            _conv_amt = _asn_settled_amount_map.get(_oid)
            if _cxp_amt is not None and _conv_amt is not None:
                try:
                    # difference = Converge - CXP
                    # negative → Converge under-settled (collected less than order total)
                    # positive → Converge over-settled (collected more than order total)
                    _diff = float(_conv_amt) - float(_cxp_amt)
                    if abs(_diff) > 0.01:
                        full_amount_mismatches.append({
                            "order_id": _oid,
                            "order_total": float(_cxp_amt),
                            "settled_amount": float(_conv_amt),
                            "difference": _diff
                        })
                except (TypeError, ValueError):
                    pass

        # IDs already flagged ACTION_REQUIRED by the classifier (date-window orders)
        _classifier_action_ids = set(classification.get("action_required_orders", []))
        # Extra mismatch orders from ASN scope that the classifier didn't see
        # (ordered outside the CXP date window but shipped within it)
        _extra_mismatch_ids = [
            m["order_id"] for m in full_amount_mismatches
            if m["order_id"] not in _classifier_action_ids
        ]

        def wr(row, values):
            for col, val in enumerate(values, start=1):
                sheet.cell(row=row, column=col).value = val

        def hdr(row, title):
            sheet.cell(row=row, column=1).value = title
            fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)
            for col in range(1, 5):
                cell = sheet.cell(row=row, column=col)
                cell.fill = fill
                cell.font = font
            sheet.merge_cells(
                start_row=row, start_column=1,
                end_row=row,   end_column=4
            )
            sheet.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            return row + 1

        def kv(row, key, value):
            sheet.cell(row=row, column=1).value = key
            sheet.cell(row=row, column=2).value = value
            sheet.cell(row=row, column=1).font  = Font(bold=True)
            return row + 1

        # ── A: Run Info ───────────────────────────────────────────────────────
        r = hdr(r, "Reconciliation Run Info")
        r = kv(r, "Start Date",      start_date)
        r = kv(r, "End Date",        end_date)
        r = kv(r, "Run Timestamp",   datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        r += 1

        # ── B: DB Query Counts ────────────────────────────────────────────────
        r = hdr(r, "Database Query Results")
        r = kv(r, "Q1 — Sales Orders fetched",              len(sales_orders))
        r = kv(r, "Q2 — Order Items fetched",               len(order_items))
        r = kv(r, "Q3 — ASN Process Numbers fetched",       len(asn_process_numbers))
        r = kv(r, "Q4 — ASN Order Totals fetched (from DB)",len(asn_order_totals))

        asn_with_totals = sum(1 for i in asn_order_totals if i.get("order_total") is not None)
        asn_pct = (
            f"{asn_with_totals}/{len(asn_process_numbers)} "
            f"({asn_with_totals / len(asn_process_numbers) * 100:.1f}%)"
            if asn_process_numbers else "0/0"
        )
        r = kv(r, "ASN Orders with CXP Total", asn_pct)

        orders_with_totals = sum(1 for i in order_totals if i.get("order_total") is not None)
        pct = f"{orders_with_totals / len(sales_orders) * 100:.1f}%" if sales_orders else "0%"
        r = kv(r, "All Orders Total Coverage",
               f"{orders_with_totals}/{len(sales_orders)} ({pct})")
        r += 1

        # ── C: Converge Batch Stats ───────────────────────────────────────────
        r = hdr(r, "Converge Batch Stats")
        cc = converge_current_result.get("stats", {})
        r = kv(r, "CURRENT — unique invoices",       cc.get("total_invoices", 0))
        r = kv(r, "CURRENT — data inconsistencies",  cc.get("data_inconsistencies", 0))
        cs = converge_settled_result.get("stats", {})
        total_inv      = cs.get("total_invoices", 0)
        settled_cnt    = cs.get("settled_count", 0)
        return_only    = cs.get("return_only_count", 0)
        anomaly_cnt    = cs.get("anomaly_count", 0)
        multi_sale_cnt = cs.get("multiple_sales_count", 0)

        r = kv(r, "SETTLED — unique invoices",
               f"{total_inv}  (= {settled_cnt} with SALE  +  {return_only} RETURN-only)")
        r = kv(r, "SETTLED — invoices with SALE (settled = True)",   settled_cnt)
        r = kv(r, "SETTLED — RETURN-only invoices (normal — refunds for prior orders)", return_only)
        r = kv(r, "SETTLED — anomaly count (MULTIPLE_SALES or NON_STANDARD only)", anomaly_cnt)
        r = kv(r, "SETTLED — multiple SALE rows", multi_sale_cnt)
        r = kv(r, "What is a Settlement Anomaly?",
               "MULTIPLE_SALES: same invoice has 2+ SALE rows in Converge (duplicate charge risk). "
               "NON_STANDARD_TYPES: unexpected transaction type. "
               "RETURN-only invoices are NOT anomalies — normal refund for an earlier order.")
        r += 1

        # ── D: Classification Summary ─────────────────────────────────────────
        r = hdr(r, "Classification Summary")
        r = kv(r, "Total Orders",                          stats.get("total", 0))
        r = kv(r, "Success — Shipped + Settled",           stats.get("success_shipped_settled", 0))
        r = kv(r, "Success — Ordered/Claimed + Approved",  stats.get("success_ordered_approved", 0))
        r = kv(r, "Success — Data inconsistency (noted)",  stats.get("verification_needed", 0))
        r = kv(r,"Rejected Order",                         stats.get("order_rejected", 0))
        r = kv(r, "Failed — Declined",                     stats.get("declined", 0))
        r = kv(r, "Failed — Fraud",                        stats.get("fraud", 0))
        r = kv(r, "Failed — Payment Cancelled",            stats.get("payment_cancelled", 0))
        r = kv(r, "Failed — Other",                        stats.get("failed_other", 0))
        r = kv(r, "Failed — Other",                        stats.get("failed_other", 0))
        r = kv(r, "Action Required — CXP Error State",     stats.get("cxp_error_state", 0))
        r = kv(r, "Action Required — ASN Not Settled",     stats.get("asn_not_settled", 0))
        r = kv(r, "Action Required — Shipped Not Settled", stats.get("shipped_not_settled", 0))
        r = kv(r, "Action Required — No Payment Data",     stats.get("no_payment_data", 0))
        r = kv(r, "Action Required — Payment Success Order Failed",
                  stats.get("payment_success_order_failed", 0))
        r = kv(r, "Action Required — Settlement Anomaly",  stats.get("settlement_anomaly", 0))
        r = kv(r, "Action Required — Amount Mismatch",     len(full_amount_mismatches))
        r = kv(r, "Action Required — Settled but No ASN",  stats.get("settled_no_asn", 0))
        r = kv(r, "Retry Successes",                       len(classification.get("retry_success_orders", [])))
        r += 1

        # ── E: Action Required Detail ─────────────────────────────────────────
        action_ids = classification.get("action_required_orders", [])

        # Augment with ASN-scope mismatches not seen by the classifier
        # (orders placed outside the CXP date window but shipped within it)
        augmented_action_ids = list(action_ids) + _extra_mismatch_ids
        total_action_count   = len(augmented_action_ids)

        r = hdr(r, f"Action Required Orders ({total_action_count})")

        if augmented_action_ids:
            wr(r, ["Order ID", "Reason", "CXP DB Status", "CXP Status",
                   "Converge Status", "Settled", "ASN", "Order Total", "Settled Amount"])
            self._apply_sub_header_style(sheet, r, 9)
            r += 1

            reason_colours = {
                "CXP_ERROR_STATE":              "FFC7CE",
                "ASN_NOT_SETTLED":              "FFDCB2",
                "SHIPPED_NOT_SETTLED":          "FFDCB2",
                "NO_PAYMENT_DATA":              "FFEB9C",
                "PAYMENT_SUCCESS_ORDER_FAILED": "DAEEF3",
                "SETTLEMENT_ANOMALY":           "FFDCB2",
                "SETTLEMENT_AMOUNT_MISMATCH":   "FFC7CE",
            }

            # Build a quick lookup for extra mismatch details
            _extra_mismatch_map = {m["order_id"]: m for m in full_amount_mismatches}

            for oid in augmented_action_ids:
                if oid in _extra_mismatch_map and oid not in _classifier_action_ids:
                    # ASN-scope mismatch order — not in orders_dict, build from mismatch data
                    m = _extra_mismatch_map[oid]
                    # Get order_state from asn_order_totals (now includes order_state column)
                    cxp_db_status = _asn_order_state_map.get(oid, "")
                    wr(r, [
                        oid,
                        "SETTLEMENT_AMOUNT_MISMATCH",
                        cxp_db_status,   # from DB via fetch_order_totals (order_state column)
                        "SHIPPED",       # must be shipped — it's in ASN
                        "NA",
                        "Yes",           # settled (it's in Converge settled)
                        "Yes",           # has ASN (that's why it's in the list)
                        f"{m['order_total']:.2f}",
                        f"{m['settled_amount']:.2f}"
                    ])
                    self._apply_row_fill(sheet, r, 9, "FFC7CE")
                else:
                    od = orders_dict.get(oid, {})
                    wr(r, [
                        oid,
                        od.get("action_reason", ""),
                        od.get("cxp_db_status", ""),
                        od.get("cxp_status", ""),
                        od.get("converge_status", "NA"),
                        "Yes" if od.get("is_settled") else "No",
                        "Yes" if od.get("has_asn") else "No",
                        od.get("order_total", ""),
                        od.get("settled_amount", "")
                    ])
                    colour = reason_colours.get(od.get("action_reason", ""), "FFEB9C")
                    self._apply_row_fill(sheet, r, 9, colour)
                r += 1
        else:
            sheet.cell(row=r, column=1).value = "✓ No orders require action"
            sheet.cell(row=r, column=1).font  = Font(bold=True, color="375623")
            r += 1

        r += 1

        # ── F: Settlement Mismatches ──────────────────────────────────────────
        # Uses the same ASN-scoped mismatch list as Reconciliation Sheet Section 2.
        # difference = Converge - CXP:  negative = under-settled, positive = over-settled
        mismatches = full_amount_mismatches
        r = hdr(r, f"Settlement Amount Mismatches ({len(mismatches)})")

        if mismatches:
            wr(r, ["Order ID", "CXP Order Total", "Converge Settled Amount", "Difference"])
            self._apply_sub_header_style(sheet, r, 4)
            r += 1
            for m in mismatches:
                diff_val = m["difference"]   # already Converge - CXP
                sign     = "+" if diff_val > 0 else ""
                wr(r, [
                    m["order_id"],
                    f"{m['order_total']:.2f}",
                    f"{m['settled_amount']:.2f}",
                    f"{sign}{diff_val:.2f}"
                ])
                self._apply_row_fill(sheet, r, 4, "FFC7CE")
                r += 1
        else:
            sheet.cell(row=r, column=1).value = "✓ No amount mismatches"
            sheet.cell(row=r, column=1).font  = Font(bold=True, color="375623")
            r += 1

        r += 1

        # ── G: ACTION REQUIRED — Settled in Converge but NO ASN ────────────
        # These orders are in action_required_orders with reason SETTLED_NO_ASN.
        # Payment was collected but the warehouse sent no ASN — confirm shipment.
        settled_no_asn_ids = classification.get("settled_no_asn_orders", [])
        r = hdr(r, f"ACTION REQUIRED — Settled but NO ASN Record ({len(settled_no_asn_ids)})")

        if settled_no_asn_ids:
            wr(r, ["Order ID", "CXP Status", "Converge Status", "Warning"])
            self._apply_sub_header_style(sheet, r, 4)
            r += 1
            orders_dict_logs = classification.get("orders", {})
            for oid in settled_no_asn_ids:
                od = orders_dict_logs.get(oid, {})
                wr(r, [
                    oid,
                    od.get("cxp_status", ""),
                    od.get("converge_status", "NA"),
                    "Payment settled but no ASN record — verify order was physically shipped"
                ])
                self._apply_row_fill(sheet, r, 4, "FFEB9C")
                r += 1
        else:
            sheet.cell(row=r, column=1).value = "✓ All settled orders have a matching ASN record — no action needed"
            sheet.cell(row=r, column=1).font  = Font(bold=True, color="375623")
            r += 1

        r += 1

        # ── H: Retry Successes ────────────────────────────────────────────────
        retry_ids = classification.get("retry_success_orders", [])
        r = hdr(r, f"Retry Successes ({len(retry_ids)})")

        if retry_ids:
            wr(r, ["Success Order ID", "Previous Failed Order ID", "Customer Email"])
            self._apply_sub_header_style(sheet, r, 3)
            r += 1
            for oid in retry_ids:
                od = orders_dict.get(oid, {})
                wr(r, [
                    oid,
                    od.get("previous_failed_attempt", ""),
                    od.get("email", "")
                ])
                self._apply_row_fill(sheet, r, 3, "C6EFCE")
                r += 1
        else:
            sheet.cell(row=r, column=1).value = "No retry successes"
            r += 1

        self._apply_borders(sheet)
        self._auto_fit_columns(sheet)

    # ================= HELPER METHODS =================
    def _apply_borders(self, sheet):
        """Apply thin border to every non-merged cell."""
        from openpyxl.styles import Border, Side
        thin   = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        from openpyxl.cell.cell import MergedCell as MC
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MC):
                    continue
                cell.border = border

    def _apply_header_style(self, sheet, row, col_count):
        dark_blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        for col in range(1, col_count + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = dark_blue_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _apply_sub_header_style(self, sheet, row, col_count):
        light_blue_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        bold_font = Font(bold=True)
        for col in range(1, col_count + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = light_blue_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _apply_row_fill(self, sheet, row, col_count, color):
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, col_count + 1):
            sheet.cell(row=row, column=col).fill = fill

    def _add_section_header(self, sheet, row, title):
        """Legacy helper — not used inside reconciliation sheet."""
        sheet.cell(row=row, column=1).value = title
        self._apply_header_style(sheet, row, 6)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        return row + 1

    def _auto_fit_columns(self, sheet):
        for column in sheet.columns:
            max_length    = 0
            column_letter = None
            for cell in column:
                if isinstance(cell, MergedCell):
                    continue
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            if column_letter:
                sheet.column_dimensions[column_letter].width = min(max_length + 3, 55)

    # ================= SAVE =================
    def save(self, directory: str = "."):
        path = f"{directory}/{self.get_filename()}"
        self.workbook.save(path)
        return path

    def to_bytes(self) -> BytesIO:
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output